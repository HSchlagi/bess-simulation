"""
Export-Funktionen für BESS-Simulation
Unterstützt PDF, Excel, CSV und Batch-Export
"""

import os
import json
import csv
import zipfile
from datetime import datetime
from io import BytesIO, StringIO
from typing import Dict, List, Any, Optional

# PDF Export
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("Warnung: reportlab nicht installiert. PDF-Export nicht verfügbar.")

# Excel Export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference, BarChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warnung: openpyxl nicht installiert. Excel-Export nicht verfügbar.")

# Pandas für Datenverarbeitung
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warnung: pandas nicht installiert. Einige Export-Funktionen nicht verfügbar.")


class BESSExporter:
    """Hauptklasse für alle Export-Funktionen"""
    
    def __init__(self):
        # Verwende absoluten Pfad vom Projektverzeichnis
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.templates_dir = os.path.join(base_dir, "export_templates")
        self.ensure_templates_dir()
    
    def ensure_templates_dir(self):
        """Erstellt das Templates-Verzeichnis falls es nicht existiert"""
        if not os.path.exists(self.templates_dir):
            os.makedirs(self.templates_dir)
    
    def export_project_pdf(self, project_data: Dict[str, Any], output_path: str = None) -> str:
        """
        Exportiert ein Projekt als PDF-Bericht
        
        Args:
            project_data: Projekt-Daten Dictionary
            output_path: Optionaler Ausgabepfad
            
        Returns:
            Pfad zur erstellten PDF-Datei
        """
        if not PDF_AVAILABLE:
            raise ImportError("PDF-Export nicht verfügbar. Installieren Sie reportlab: pip install reportlab")
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Verwende absoluten Pfad vom Projektverzeichnis
            base_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(base_dir, "export", f"project_{project_data.get('id', 'unknown')}_{timestamp}.pdf")
        
        # Erstelle Ausgabeverzeichnis
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # PDF erstellen
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkgreen
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        # Titel
        story.append(Paragraph("BESS-Projekt Bericht", title_style))
        story.append(Spacer(1, 20))
        
        # Projekt-Informationen
        story.append(Paragraph("Projekt-Übersicht", heading_style))
        
        project_info = [
            ["Projekt-Name:", project_data.get('name', 'N/A')],
            ["Standort:", project_data.get('location', 'N/A')],
            ["BESS-Größe:", f"{project_data.get('bess_size', 0):.1f} kWh"],
            ["BESS-Leistung:", f"{project_data.get('bess_power', 0):.1f} kW"],
            ["PV-Leistung:", f"{project_data.get('pv_power', 0):.1f} kW"],
            ["Wasserkraft:", f"{project_data.get('hydro_power', 0):.1f} kW"],
            ["Stromkosten:", f"{project_data.get('current_electricity_cost', 0):.3f} €/kWh"],
        ]
        
        project_table = Table(project_info, colWidths=[2*inch, 4*inch])
        project_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(project_table)
        story.append(Spacer(1, 20))
        
        # Wirtschaftlichkeitsanalyse
        if 'economic_data' in project_data:
            story.append(Paragraph("Wirtschaftlichkeitsanalyse", heading_style))
            
            econ_data = project_data['economic_data']
            econ_info = [
                ["Gesamtkosten:", f"{econ_data.get('total_cost', 0):,.2f} €"],
                ["Jährliche Einsparungen:", f"{econ_data.get('annual_savings', 0):,.2f} €"],
                ["Amortisationszeit:", f"{econ_data.get('payback_period', 0):.1f} Jahre"],
                ["ROI:", f"{econ_data.get('roi', 0):.1f}%"],
            ]
            
            econ_table = Table(econ_info, colWidths=[2*inch, 4*inch])
            econ_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(econ_table)
            story.append(Spacer(1, 20))
        
        # Technische Details
        story.append(Paragraph("Technische Spezifikationen", heading_style))
        
        tech_info = [
            ["Batterietyp:", project_data.get('battery_type', 'Lithium-Ion')],
            ["Entladetiefe:", f"{project_data.get('depth_of_discharge', 80):.0f}%"],
            ["Lebensdauer:", f"{project_data.get('battery_lifetime', 10):.0f} Jahre"],
            ["Wirkungsgrad:", f"{project_data.get('efficiency', 90):.1f}%"],
        ]
        
        tech_table = Table(tech_info, colWidths=[2*inch, 4*inch])
        tech_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(tech_table)
        story.append(Spacer(1, 20))
        
        # Footer
        story.append(Paragraph(f"Bericht erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 
                              styles['Normal']))
        story.append(Paragraph("BESS-Simulation System", styles['Normal']))
        
        # PDF generieren
        doc.build(story)
        return output_path
    
    def export_simulation_excel(self, simulation_data: Dict[str, Any], output_path: str = None) -> str:
        """
        Exportiert Simulationsdaten als Excel-Datei
        
        Args:
            simulation_data: Simulations-Daten Dictionary
            output_path: Optionaler Ausgabepfad
            
        Returns:
            Pfad zur erstellten Excel-Datei
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel-Export nicht verfügbar. Installieren Sie openpyxl: pip install openpyxl")
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Verwende absoluten Pfad vom Projektverzeichnis
            base_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(base_dir, "export", f"simulation_{timestamp}.xlsx")
        
        # Erstelle Ausgabeverzeichnis
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Excel-Workbook erstellen
        wb = openpyxl.Workbook()
        
        # Styles definieren
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Projekt-Übersicht
        ws_overview = wb.active
        ws_overview.title = "Projekt-Übersicht"
        
        # Überschrift
        ws_overview['A1'] = "BESS-Simulation - Projektübersicht"
        ws_overview['A1'].font = Font(size=16, bold=True, color="366092")
        ws_overview.merge_cells('A1:D1')
        
        # Projekt-Daten
        project_data = simulation_data.get('project', {})
        overview_data = [
            ["Projekt-Name", project_data.get('name', 'N/A')],
            ["Standort", project_data.get('location', 'N/A')],
            ["BESS-Größe (kWh)", project_data.get('bess_size', 0)],
            ["BESS-Leistung (kW)", project_data.get('bess_power', 0)],
            ["PV-Leistung (kW)", project_data.get('pv_power', 0)],
            ["Wasserkraft (kW)", project_data.get('hydro_power', 0)],
            ["Stromkosten (€/kWh)", project_data.get('current_electricity_cost', 0)],
        ]
        
        for i, (label, value) in enumerate(overview_data, start=3):
            ws_overview[f'A{i}'] = label
            ws_overview[f'B{i}'] = value
            ws_overview[f'A{i}'].font = Font(bold=True)
        
        # Zeitreihen-Daten
        if 'time_series' in simulation_data:
            ws_timeseries = wb.create_sheet("Zeitreihen")
            
            # Header
            headers = ["Datum", "Zeit", "Last (kW)", "PV-Erzeugung (kW)", "Batterie-Ladung (kW)", 
                      "Batterie-Entladung (kW)", "Netzbezug (kW)", "Netzeinspeisung (kW)"]
            
            for col, header in enumerate(headers, 1):
                cell = ws_timeseries.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Daten einfügen
            time_series = simulation_data['time_series']
            for row, data_point in enumerate(time_series, 2):
                ws_timeseries.cell(row=row, column=1, value=data_point.get('date', ''))
                ws_timeseries.cell(row=row, column=2, value=data_point.get('time', ''))
                ws_timeseries.cell(row=row, column=3, value=data_point.get('load', 0))
                ws_timeseries.cell(row=row, column=4, value=data_point.get('pv_generation', 0))
                ws_timeseries.cell(row=row, column=5, value=data_point.get('battery_charge', 0))
                ws_timeseries.cell(row=row, column=6, value=data_point.get('battery_discharge', 0))
                ws_timeseries.cell(row=row, column=7, value=data_point.get('grid_import', 0))
                ws_timeseries.cell(row=row, column=8, value=data_point.get('grid_export', 0))
            
            # Spaltenbreite anpassen
            for column in ws_timeseries.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_timeseries.column_dimensions[column_letter].width = adjusted_width
        
        # Wirtschaftlichkeitsanalyse
        if 'economic_analysis' in simulation_data:
            ws_economic = wb.create_sheet("Wirtschaftlichkeit")
            
            # Header
            econ_headers = ["Kennzahl", "Wert", "Einheit"]
            for col, header in enumerate(econ_headers, 1):
                cell = ws_economic.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Wirtschaftlichkeitsdaten
            econ_data = simulation_data['economic_analysis']
            econ_rows = [
                ["Gesamtkosten", econ_data.get('total_cost', 0), "€"],
                ["Jährliche Einsparungen", econ_data.get('annual_savings', 0), "€"],
                ["Amortisationszeit", econ_data.get('payback_period', 0), "Jahre"],
                ["ROI", econ_data.get('roi', 0), "%"],
                ["NPV", econ_data.get('npv', 0), "€"],
                ["IRR", econ_data.get('irr', 0), "%"],
            ]
            
            for row, (label, value, unit) in enumerate(econ_rows, 2):
                ws_economic.cell(row=row, column=1, value=label)
                ws_economic.cell(row=row, column=2, value=value)
                ws_economic.cell(row=row, column=3, value=unit)
                ws_economic.cell(row=row, column=1).font = Font(bold=True)
        
        # Speichern
        wb.save(output_path)
        return output_path
    
    def export_data_csv(self, data: List[Dict[str, Any]], output_path: str = None, 
                       fieldnames: List[str] = None) -> str:
        """
        Exportiert Daten als CSV-Datei
        
        Args:
            data: Liste von Daten-Dictionaries
            output_path: Optionaler Ausgabepfad
            fieldnames: Optional - Spaltennamen
            
        Returns:
            Pfad zur erstellten CSV-Datei
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Verwende absoluten Pfad vom Projektverzeichnis
            base_dir = os.path.dirname(os.path.abspath(__file__))
            output_path = os.path.join(base_dir, "export", f"data_{timestamp}.csv")
        
        # Erstelle Ausgabeverzeichnis
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # CSV schreiben
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            if fieldnames is None and data:
                fieldnames = list(data[0].keys())
            
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        
        return output_path
    
    def batch_export_projects(self, projects: List[Dict[str, Any]], 
                            export_formats: List[str] = None) -> str:
        """
        Exportiert mehrere Projekte in verschiedenen Formaten
        
        Args:
            projects: Liste von Projekt-Daten
            export_formats: Liste von Export-Formaten ['pdf', 'excel', 'csv']
            
        Returns:
            Pfad zur ZIP-Datei mit allen Exports
        """
        if export_formats is None:
            export_formats = ['pdf', 'excel']
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Verwende absoluten Pfad vom Projektverzeichnis
        base_dir = os.path.dirname(os.path.abspath(__file__))
        zip_path = os.path.join(base_dir, "export", f"batch_export_{timestamp}.zip")
        
        # Erstelle Ausgabeverzeichnis
        os.makedirs(os.path.dirname(zip_path), exist_ok=True)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for i, project in enumerate(projects):
                project_name = project.get('name', f'project_{i}').replace(' ', '_')
                
                # PDF Export
                if 'pdf' in export_formats and PDF_AVAILABLE:
                    try:
                        pdf_path = self.export_project_pdf(project)
                        zipf.write(pdf_path, f"{project_name}/{os.path.basename(pdf_path)}")
                        os.remove(pdf_path)  # Temporäre Datei löschen
                    except Exception as e:
                        print(f"PDF-Export fehlgeschlagen für {project_name}: {e}")
                
                # Excel Export
                if 'excel' in export_formats and EXCEL_AVAILABLE:
                    try:
                        excel_path = self.export_simulation_excel({'project': project})
                        zipf.write(excel_path, f"{project_name}/{os.path.basename(excel_path)}")
                        os.remove(excel_path)  # Temporäre Datei löschen
                    except Exception as e:
                        print(f"Excel-Export fehlgeschlagen für {project_name}: {e}")
                
                # CSV Export
                if 'csv' in export_formats:
                    try:
                        # Projekt-Daten als CSV
                        csv_data = [project]
                        csv_path = self.export_data_csv(csv_data)
                        zipf.write(csv_path, f"{project_name}/{os.path.basename(csv_path)}")
                        os.remove(csv_path)  # Temporäre Datei löschen
                    except Exception as e:
                        print(f"CSV-Export fehlgeschlagen für {project_name}: {e}")
        
        return zip_path
    
    def create_export_template(self, template_name: str, template_data: Dict[str, Any]) -> str:
        """
        Erstellt ein Export-Template
        
        Args:
            template_name: Name des Templates
            template_data: Template-Daten
            
        Returns:
            Pfad zum gespeicherten Template
        """
        template_path = os.path.join(self.templates_dir, f"{template_name}.json")
        
        with open(template_path, 'w', encoding='utf-8') as f:
            json.dump(template_data, f, indent=2, ensure_ascii=False)
        
        return template_path
    
    def load_export_template(self, template_name: str) -> Dict[str, Any]:
        """
        Lädt ein Export-Template
        
        Args:
            template_name: Name des Templates
            
        Returns:
            Template-Daten
        """
        template_path = os.path.join(self.templates_dir, f"{template_name}.json")
        
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template '{template_name}' nicht gefunden")
        
        with open(template_path, 'r', encoding='utf-8') as f:
            return json.load(f)


# Vordefinierte Templates
DEFAULT_TEMPLATES = {
    "kundenbericht": {
        "name": "Kundenbericht",
        "sections": ["projekt_uebersicht", "wirtschaftlichkeit", "technische_details"],
        "include_charts": True,
        "language": "de"
    },
    "technische_dokumentation": {
        "name": "Technische Dokumentation",
        "sections": ["technische_details", "simulationsergebnisse", "zeitreihen"],
        "include_charts": True,
        "language": "de"
    },
    "wirtschaftlichkeitsanalyse": {
        "name": "Wirtschaftlichkeitsanalyse",
        "sections": ["wirtschaftlichkeit", "szenarien_vergleich", "sensitivity_analysis"],
        "include_charts": True,
        "language": "de"
    }
}


def create_default_templates():
    """Erstellt die Standard-Templates"""
    exporter = BESSExporter()
    
    for template_id, template_data in DEFAULT_TEMPLATES.items():
        try:
            exporter.create_export_template(template_id, template_data)
            print(f"Template '{template_id}' erstellt")
        except Exception as e:
            print(f"Fehler beim Erstellen von Template '{template_id}': {e}")


if __name__ == "__main__":
    # Erstelle Standard-Templates
    create_default_templates()
    print("Export-Funktionen erfolgreich initialisiert!")
