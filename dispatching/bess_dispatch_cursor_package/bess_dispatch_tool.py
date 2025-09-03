#!/usr/bin/env python3
"""
bess_dispatch_tool.py – BESS Dispatch, SoC-Simulation, Charts & CSVs
Features:
- Load Excel base sheet (hour/price/dispatch)
- Ensure time axis (hourly or --freq_minutes incl. 15-min) & multi-day replication
- Recompute SoC simulation with constraints & efficiencies
- Settlement (revenues/costs/cashflow)
- Plots: SoC, cumulative cashflow, price, feasible dispatch
- Redispatch: apply calls (delta/absolute) with compensation; incremental report vs baseline
- Optional: write Redispatch sheets into the Excel file
"""

import argparse
from datetime import datetime
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt



# ---------- Country profiles & compensation ----------

COUNTRY_PROFILES = {
    "DE": {
        "default_freq_minutes": 15,
        "notes": "Germany Redispatch 2.0; standard 15-min slots; data exchange via Connect+/RAIDA.",
        "default_rd_comp_mode": "market_price"  # use base price as reference
    },
    "AT": {
        "default_freq_minutes": 15,
        "notes": "Austria APG redispatch/countertrading; 15-min scheduling.",
        "default_rd_comp_mode": "flat_csv"  # expect explicit compensation in CSV
    }
}

def compute_compensation(series_delta_e_out, series_delta_e_in, base_price, comp_price_series, mode: str, premium: float = 0.0):
    """
    Returns per-slot compensation in EUR according to mode:
      - 'flat_csv': use comp_price_series [EUR/MWh]
      - 'market_price': use base_price [EUR/MWh]
      - 'premium': base_price + premium [EUR/MWh]
    Compensation volume uses |ΔE_out| + |ΔE_in| (both grid-side).
    """
    import numpy as np
    vol = series_delta_e_out.abs().to_numpy() + series_delta_e_in.abs().to_numpy()
    if mode == "flat_csv":
        price = comp_price_series.to_numpy()
    elif mode == "premium":
        price = base_price.to_numpy() + premium
    else:  # market_price
        price = base_price.to_numpy()
    return vol * price
# ---------- Base helpers ----------

def load_parameters_from_sheet(xlsx_path: Path) -> dict:
    try:
        params = pd.read_excel(xlsx_path, sheet_name="Parameter")
        p = {row["Parameter"]: row["Wert"] for _, row in params.iterrows()}
    except Exception:
        p = {}
    defaults = {
        "Wirkungsgrad Entladen": 0.92,
        "Wirkungsgrad Laden": 0.92,
        "Zeitschrittdauer [h]": 1.0,
        "Kapazität [MWh]": 8.0,
        "P_max_Entladen [MW]": 2.0,
        "P_max_Laden [MW]": 2.0,
        "SoC_init [%]": 50.0,
        "SoC_min [%]": 5.0,
        "SoC_max [%]": 95.0,
    }
    for k, v in defaults.items():
        p.setdefault(k, v)
    return p


def load_base_sheet(xlsx_path: Path) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=0)
    rename = {
        "Stunde": "hour",
        "Strompreis_[EUR/MWh]": "price_eur_mwh",
        "BESS_Dispatch_[MW]": "dispatch_mw",
    }
    for k in rename:
        if k not in df.columns:
            raise ValueError(f"Spalte '{k}' fehlt im Basissheet!")
    df = df.rename(columns=rename)
    df = df[["hour", "price_eur_mwh", "dispatch_mw"]].copy()
    return df


def ensure_time_index(base_df, freq_minutes=60, start_ts=None, days=None):
    """
    Ensure a time axis. If Excel provides 'timestamp' or 'Zeit', use it.
    Otherwise synthesize timestamps from 'hour' and replicate across days if needed.
    If freq_minutes != 60, upsample to higher frequency (forward-fill within hour).
    Returns DataFrame with 'timestamp' and integer 'slot' columns.
    """
    df = base_df.copy()
    if 'timestamp' in df.columns:
        ts = pd.to_datetime(df['timestamp'])
    elif 'Zeit' in df.columns:
        ts = pd.to_datetime(df['Zeit'])
    else:
        # infer days
        if days is None:
            n = len(df)
            days = max(n // 24, 1) if (n % 24 == 0 and n >= 24) else 1
        if start_ts is None:
            start_ts = pd.Timestamp('2025-01-01 00:00')
        if len(df) == 24 and days > 1:
            df = pd.concat([df]*days, ignore_index=True)
        n = len(df)
        ts = pd.date_range(start=start_ts, periods=n, freq='H')
    df['timestamp'] = ts

    if freq_minutes and int(freq_minutes) != 60:
        freq = f'{int(freq_minutes)}T'
        df = df.set_index('timestamp')
        start = df.index.min()
        end = df.index.max() + pd.Timedelta(hours=1) - pd.Timedelta(minutes=freq_minutes)
        target_index = pd.date_range(start=start, end=end, freq=freq)
        df = df.reindex(df.index.union(target_index)).sort_index()
        df[['price_eur_mwh','dispatch_mw']] = df[['price_eur_mwh','dispatch_mw']].ffill()
        df = df.loc[target_index]
        df = df.reset_index().rename(columns={'index':'timestamp'})
    else:
        df = df.reset_index(drop=True)

    df['slot'] = range(len(df))
    return df


def simulate_soc(base: pd.DataFrame, params: dict) -> pd.DataFrame:
    eta_dis = float(params["Wirkungsgrad Entladen"])
    eta_cha = float(params["Wirkungsgrad Laden"])
    dt = float(params["Zeitschrittdauer [h]"])
    cap = float(params["Kapazität [MWh]"])
    pmax_dis = float(params["P_max_Entladen [MW]"])
    pmax_cha = float(params["P_max_Laden [MW]"])
    soc_init_pct = float(params["SoC_init [%]"])
    soc_min_pct = float(params["SoC_min [%]"])
    soc_max_pct = float(params["SoC_max [%]"])

    soc_min = cap * soc_min_pct / 100.0
    soc_max = cap * soc_max_pct / 100.0
    soc = cap * soc_init_pct / 100.0

    hours = base["hour"].to_list() if "hour" in base.columns else list(range(len(base)))
    price = base["price_eur_mwh"].to_numpy()
    dispatch_plan = base["dispatch_mw"].to_numpy()

    dispatch_limited, dispatch_feasible = [], []
    e_out_grid, e_in_grid = [], []
    soc_list, soc_pct_list, flags = [], [], []

    for t, d_plan in enumerate(dispatch_plan):
        d_lim = min(max(d_plan, -pmax_cha), pmax_dis)
        d_feas = d_lim
        reason = "OK"
        if d_lim >= 0:
            e_out = d_lim * dt
            e_batt = e_out / eta_dis if eta_dis > 0 else 1e18
            if soc - e_batt < soc_min - 1e-12:
                avail = max(soc - soc_min, 0.0)
                e_out_max = avail * eta_dis
                d_feas = max(e_out_max / dt, 0.0)
                reason = "Clip: SoC_min"
            e_out = d_feas * dt
            e_batt = e_out / eta_dis if eta_dis > 0 else 0.0
            soc = max(soc - e_batt, soc_min)
            e_in = 0.0
        else:
            e_in = (-d_lim) * dt
            e_batt = e_in * eta_cha
            if soc + e_batt > soc_max + 1e-12:
                room = max(soc_max - soc, 0.0)
                e_in_max = room / eta_cha if eta_cha > 0 else 0.0
                d_feas = -max(e_in_max / dt, 0.0)
                reason = "Clip: SoC_max"
            e_in = (-d_feas) * dt
            e_batt = e_in * eta_cha
            soc = min(soc + e_batt, soc_max)
            e_out = 0.0

        dispatch_limited.append(d_lim)
        dispatch_feasible.append(d_feas)
        e_out_grid.append(e_out)
        e_in_grid.append(e_in)
        soc_list.append(soc)
        soc_pct_list.append(100.0 * soc / cap if cap > 0 else 0.0)

        if reason == "OK" and abs(d_lim - d_plan) > 1e-12:
            reason = "Clip: P_max"
        flags.append(reason)

    sim = pd.DataFrame({
        "Stunde": hours,
        "Preis_[EUR/MWh]": price,
        "Dispatch_Plan_[MW]": dispatch_plan,
        "Dispatch_Limit_[MW]": dispatch_limited,
        "Dispatch_Feasible_[MW]": dispatch_feasible,
        "E_out_to_Grid_[MWh]": e_out_grid,
        "E_in_from_Grid_[MWh]": e_in_grid,
        "SoC_[MWh]": soc_list,
        "SoC_[%]": soc_pct_list,
        "Check": flags
    })
    if "timestamp" in base.columns:
        sim["timestamp"] = base["timestamp"]
    return sim


def settlement_from_sim(sim_df: pd.DataFrame) -> pd.DataFrame:
    price = sim_df["Preis_[EUR/MWh]"].to_numpy()
    e_out = sim_df["E_out_to_Grid_[MWh]"].to_numpy()
    e_in = sim_df["E_in_from_Grid_[MWh]"].to_numpy()
    revenue = e_out * price
    costs = e_in * price
    cash = revenue - costs
    df = pd.DataFrame({
        "Stunde": sim_df["Stunde"].to_numpy(),
        "Preis_[EUR/MWh]": price,
        "E_out_to_Grid_[MWh]": e_out,
        "E_in_from_Grid_[MWh]": e_in,
        "Einnahmen_[EUR]": revenue,
        "Kosten_[EUR]": costs,
        "Cashflow_[EUR]": cash
    })
    df["Kumuliert_[EUR]"] = df["Cashflow_[EUR]"].cumsum()
    if "timestamp" in sim_df.columns:
        df["timestamp"] = sim_df["timestamp"]
    return df


def plot_soc(sim_df: pd.DataFrame, outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    fig_path = outdir / "SoC_Verlauf.png"
    plt.figure(figsize=(10, 5))
    x = sim_df['timestamp'] if 'timestamp' in sim_df.columns else sim_df['Stunde']
    plt.plot(x, sim_df["SoC_[%]"], linewidth=2)
    plt.title("SoC-Verlauf [%]")
    plt.xlabel("Zeit" if 'timestamp' in sim_df.columns else "Stunde")
    plt.ylabel("SoC [%]")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(fig_path, dpi=150)
    plt.close()
    return fig_path


def plot_cashflow(ab_df: pd.DataFrame, outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    fig_path = outdir / "Cashflow_Kumuliert.png"
    plt.figure(figsize=(10, 5))
    x = ab_df['timestamp'] if 'timestamp' in ab_df.columns else ab_df['Stunde']
    plt.plot(x, ab_df["Kumuliert_[EUR]"], linewidth=2)
    plt.title("Kumulierter Cashflow [EUR]")
    plt.xlabel("Zeit" if 'timestamp' in ab_df.columns else "Stunde")
    plt.ylabel("Kumuliert [EUR]")
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(fig_path, dpi=150)
    plt.close()
    return fig_path


def plot_price_and_dispatch(base: pd.DataFrame, sim_df: pd.DataFrame, outdir: Path) -> (Path, Path):
    outdir.mkdir(parents=True, exist_ok=True)
    # Price
    ppath = outdir / "Preis.png"
    plt.figure(figsize=(10,5))
    x = base['timestamp'] if 'timestamp' in base.columns else base['hour']
    plt.plot(x, base["price_eur_mwh"], linewidth=2)
    plt.title("Strompreis [EUR/MWh]")
    plt.xlabel("Zeit" if 'timestamp' in base.columns else "Stunde")
    plt.ylabel("Preis [EUR/MWh]")
    plt.grid(True, alpha=0.3)
    plt.tight_layout(); plt.savefig(ppath, dpi=150); plt.close()

    # Dispatch feasible
    dpath = outdir / "Dispatch_Feasible.png"
    plt.figure(figsize=(10,5))
    x2 = sim_df['timestamp'] if 'timestamp' in sim_df.columns else sim_df['Stunde']
    plt.bar(x2, sim_df["Dispatch_Feasible_[MW]"], alpha=0.7)
    plt.title("BESS Dispatch (realisierbar) [MW]")
    plt.xlabel("Zeit" if 'timestamp' in sim_df.columns else "Stunde")
    plt.ylabel("MW")
    plt.grid(True, axis="y", alpha=0.3)
    plt.tight_layout(); plt.savefig(dpath, dpi=150); plt.close()
    return ppath, dpath


def write_csvs(base: pd.DataFrame, sim_df: pd.DataFrame, ab_df: pd.DataFrame, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    base.to_csv(outdir / "base_input.csv", index=False)
    sim_df.to_csv(outdir / "soc_simulation.csv", index=False)
    ab_df.to_csv(outdir / "abrechnung_soc.csv", index=False)


# ---------- Redispatch helpers ----------

def read_redispatch_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    ren = {}
    for c in df.columns:
        lc = c.lower().strip()
        if lc in ("start","start_time","start_timestamp","timestamp","zeit","begin"):
            ren[c] = "start"
        elif lc in ("hour","start_hour"):
            ren[c] = "start_hour"
        elif lc in ("slot","start_slot"):
            ren[c] = "start_slot"
        elif lc in ("duration","duration_slots","slots"):
            ren[c] = "duration_slots"
        elif lc in ("end","end_time","end_timestamp"):
            ren[c] = "end"
        elif lc in ("end_hour",):
            ren[c] = "end_hour"
        elif lc in ("end_slot",):
            ren[c] = "end_slot"
        elif lc in ("power","power_mw","p_mw","leistung_mw"):
            ren[c] = "power_mw"
        elif lc in ("mode","modus","type","typ"):
            ren[c] = "mode"
        elif lc in ("compensation_eur_mwh","comp_eur_mwh","preis_eur_mwh","price_eur_mwh"):
            ren[c] = "compensation_eur_mwh"
        elif lc in ("reason","grund","note"):
            ren[c] = "reason"
    df = df.rename(columns=ren)
    if "mode" not in df.columns:
        df["mode"] = "delta"
    if "compensation_eur_mwh" not in df.columns:
        df["compensation_eur_mwh"] = 0.0
    return df


def normalize_redispatch_calls(rd_df: pd.DataFrame, base: pd.DataFrame) -> pd.DataFrame:
    n = len(base)
    ts_map = {}
    if "timestamp" in base.columns:
        for i, ts in enumerate(base["timestamp"]):
            ts_map[pd.Timestamp(ts)] = i

    rows = []
    for _, r in rd_df.iterrows():
        power = float(r["power_mw"])
        mode = str(r.get("mode","delta")).lower()
        comp = float(r.get("compensation_eur_mwh", 0.0))
        label = str(r.get("reason",""))

        start_slot = None
        if "start_slot" in r and pd.notna(r["start_slot"]):
            start_slot = int(r["start_slot"])
        elif "start_hour" in r and pd.notna(r["start_hour"]):
            h = int(r["start_hour"])
            if "hour" in base.columns:
                idxs = list(np.where(base["hour"].to_numpy()==h)[0])
                start_slot = idxs[0] if idxs else None
        elif "start" in r and pd.notna(r["start"]):
            ts = pd.to_datetime(r["start"])
            if len(ts_map):
                start_slot = ts_map.get(ts, None)
                if start_slot is None:
                    diffs = (pd.to_datetime(base["timestamp"]) - ts).abs()
                    start_slot = int(diffs.idxmin())
            else:
                start_slot = int(np.argmin(np.abs(pd.to_datetime(base["timestamp"]) - ts))) if "timestamp" in base.columns else 0
        if start_slot is None:
            raise ValueError(f"Could not resolve start slot for redispatch row: {r.to_dict()}")

        if "duration_slots" in r and pd.notna(r["duration_slots"]):
            dur = int(r["duration_slots"])
            end_slot = start_slot + max(dur,1)
        elif "end_slot" in r and pd.notna(r["end_slot"]):
            end_slot = int(r["end_slot"])
        elif "end_hour" in r and pd.notna(r["end_hour"]):
            eh = int(r["end_hour"])
            if "hour" in base.columns:
                idxs = list(np.where(base["hour"].to_numpy()==eh)[0])
                end_slot = (idxs[-1]+1) if idxs else (start_slot+1)
            else:
                end_slot = start_slot+1
        elif "end" in r and pd.notna(r["end"]):
            te = pd.to_datetime(r["end"])
            if "timestamp" in base.columns:
                try:
                    end_slot = list(base.index[base["timestamp"]==te])[0] + 1
                except Exception:
                    diffs = (pd.to_datetime(base["timestamp"]) - te).abs()
                    end_slot = int(diffs.idxmin()) + 1
            else:
                end_slot = start_slot + 1
        else:
            end_slot = start_slot + 1

        start_slot = max(0, min(n-1, int(start_slot)))
        end_slot = max(start_slot+1, min(n, int(end_slot)))
        rows.append({
            "start_slot": start_slot,
            "end_slot": end_slot,
            "power_mw": power,
            "mode": mode,
            "compensation_eur_mwh": comp,
            "label": label
        })
    return pd.DataFrame(rows)


def apply_redispatch(base: pd.DataFrame, calls_norm: pd.DataFrame):
    adj = base.copy()
    rd_delta = np.zeros(len(base))
    comp_price = np.zeros(len(base))
    abs_mask = np.zeros(len(base), dtype=bool)
    abs_vals = np.zeros(len(base))

    for _, c in calls_norm.iterrows():
        s, e = int(c["start_slot"]), int(c["end_slot"])
        p, mode = float(c["power_mw"]), c["mode"]
        price = float(c["compensation_eur_mwh"])
        comp_price[s:e] += price
        if mode == "absolute":
            abs_mask[s:e] = True
            abs_vals[s:e] = p
        else:
            rd_delta[s:e] += p

    plan = adj["dispatch_mw"].to_numpy()
    adj_dispatch = plan + rd_delta
    adj_dispatch = np.where(abs_mask, abs_vals, adj_dispatch)
    adj["dispatch_mw"] = plan
    adj["dispatch_mw_adj"] = adj_dispatch
    return adj, pd.Series(rd_delta), pd.Series(comp_price)


def redispatch_analysis(base: pd.DataFrame, params: dict, calls_norm: pd.DataFrame):
    sim_base = simulate_soc(base[["hour","price_eur_mwh","dispatch_mw"]], params)
    ab_base = settlement_from_sim(sim_base)

    base_adj, rd_delta, comp_price = apply_redispatch(base, calls_norm)
    base_rd = base_adj[["hour","price_eur_mwh","dispatch_mw_adj"]].rename(columns={"dispatch_mw_adj":"dispatch_mw"})
    if "timestamp" in base.columns:
        base_rd["timestamp"] = base["timestamp"]
    sim_rd = simulate_soc(base_rd, params)
    ab_rd = settlement_from_sim(sim_rd)

    cols = ["E_out_to_Grid_[MWh]","E_in_from_Grid_[MWh]","Einnahmen_[EUR]","Kosten_[EUR]","Cashflow_[EUR]"]
    diff = ab_rd[cols].values - ab_base[cols].values
    inc = pd.DataFrame(diff, columns=[f"Δ{c}" for c in cols])
    inc["Preis_[EUR/MWh]"] = ab_rd["Preis_[EUR/MWh]"]
    inc["timestamp"] = ab_rd["timestamp"] if "timestamp" in ab_rd.columns else None
    inc["slot"] = range(len(inc))

    e_abs = (inc["ΔE_out_to_Grid_[MWh]"].abs() + inc["ΔE_in_from_Grid_[MWh]"].abs())
    comp = compute_compensation(inc["ΔE_out_to_Grid_[MWh]"], inc["ΔE_in_from_Grid_[MWh]"], ab_rd["Preis_[EUR/MWh]"], comp_price, args.rd_comp_mode or (COUNTRY_PROFILES.get(args.country, {}) or {}).get('default_rd_comp_mode', 'flat_csv'), premium=args.rd_premium)
    inc["Compensation_[EUR]"] = comp
    inc["ΔCashflow_with_Comp_[EUR]"] = inc["ΔCashflow_[EUR]"] + inc["Compensation_[EUR]"]
    inc["Kumuliert_ΔCF_with_Comp_[EUR]"] = inc["ΔCashflow_with_Comp_[EUR]"].cumsum()

    return sim_base, ab_base, sim_rd, ab_rd, inc, base_adj, calls_norm


# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser(description="BESS Dispatch – Excel laden, SoC-Simulation, Charts & CSVs (+Redispatch)")
    ap.add_argument("--excel", required=True, help="Pfad zur Excel-Datei (z. B. BESS_Dispatch_Beispiel.xlsx)")
    ap.add_argument("--outdir", default="./out", help="Ausgabeverzeichnis")
    ap.add_argument("--recompute", action="store_true", help="SoC-Simulation neu berechnen (statt aus Excel-Sheets lesen)")
    # Optional overrides
    ap.add_argument("--eta_dis", type=float, help="Wirkungsgrad Entladen (override)")
    ap.add_argument("--eta_cha", type=float, help="Wirkungsgrad Laden (override)")
    ap.add_argument("--dt", type=float, help="Zeitschrittdauer [h] (override)")
    ap.add_argument("--cap", type=float, help="Kapazität [MWh] (override)")
    ap.add_argument("--pmax_dis", type=float, help="P_max_Entladen [MW] (override)")
    ap.add_argument("--pmax_cha", type=float, help="P_max_Laden [MW] (override)")
    ap.add_argument("--soc_init", type=float, help="SoC_init [%] (override)")
    ap.add_argument("--soc_min", type=float, help="SoC_min [%] (override)")
    ap.add_argument("--soc_max", type=float, help="SoC_max [%] (override)")
    # Time axis & multi-day
    ap.add_argument("--freq_minutes", type=int, default=60, help="Zeitraster in Minuten (60=stündlich, 15=Viertelstunde)")
    ap.add_argument("--start", type=str, help="Start-Zeitstempel, z. B. '2025-01-01 00:00' (falls keine Timestamps vorhanden)")
    ap.add_argument("--days", type=int, help="Anzahl Tage (bei 24 Zeilen wird pattern repliziert)")
    # Redispatch
    ap.add_argument("--rd_csv", type=str, help="Pfad zur Redispatch-CSV (optional)")
    ap.add_argument("--write_excel", action="store_true", help="Schreibe Redispatch-Sheets in die Excel-Datei")
    ap.add_argument("--country", type=str, choices=["DE","AT"], help="Länderprofil: DE oder AT")
    ap.add_argument("--rd_comp_mode", type=str, choices=["flat_csv","market_price","premium"], help="Redispatch-Compensation-Modus")
    ap.add_argument("--rd_premium", type=float, default=0.0, help="Premium [EUR/MWh] für Modus 'premium'")

    args = ap.parse_args()
    # Apply country defaults (freq & compensation mode)
    if args.country:
        prof = COUNTRY_PROFILES.get(args.country)
        if prof:
            if (args.freq_minutes == 60):  # user didn't change from default
                args.freq_minutes = prof.get("default_freq_minutes", args.freq_minutes)
            if not args.rd_comp_mode:
                args.rd_comp_mode = prof.get("default_rd_comp_mode")

    xlsx_path = Path(args.excel).expanduser().resolve()
    outdir = Path(args.outdir).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    base = load_base_sheet(xlsx_path)
    params = load_parameters_from_sheet(xlsx_path)

    # Overrides
    if args.eta_dis is not None: params["Wirkungsgrad Entladen"] = args.eta_dis
    if args.eta_cha is not None: params["Wirkungsgrad Laden"] = args.eta_cha
    if args.dt is not None:
        params["Zeitschrittdauer [h]"] = args.dt
    else:
        if args.freq_minutes:
            params["Zeitschrittdauer [h]"] = float(args.freq_minutes)/60.0
    if args.cap is not None: params["Kapazität [MWh]"] = args.cap
    if args.pmax_dis is not None: params["P_max_Entladen [MW]"] = args.pmax_dis
    if args.pmax_cha is not None: params["P_max_Laden [MW]"] = args.pmax_cha
    if args.soc_init is not None: params["SoC_init [%]"] = args.soc_init
    if args.soc_min is not None: params["SoC_min [%]"] = args.soc_min
    if args.soc_max is not None: params["SoC_max [%]"] = args.soc_max

    # Time axis / resample
    start_ts = pd.to_datetime(args.start) if args.start else None
    base = ensure_time_index(base, freq_minutes=args.freq_minutes, start_ts=start_ts, days=args.days)

    # Compute or read existing sheets
    if args.recompute:
        sim_df = simulate_soc(base, params)
        ab_df = settlement_from_sim(sim_df)
    else:
        try:
            sim_df = pd.read_excel(xlsx_path, sheet_name="SoC_Simulation")
            ab_df = pd.read_excel(xlsx_path, sheet_name="Abrechnung_SoC")
            # if existing sheets are hourly but we asked for 15-min, recompute
            if args.freq_minutes and len(sim_df) != len(base):
                sim_df = simulate_soc(base, params)
                ab_df = settlement_from_sim(sim_df)
        except Exception:
            sim_df = simulate_soc(base, params)
            ab_df = settlement_from_sim(sim_df)

    # Exports base
    write_csvs(base, sim_df, ab_df, outdir)
    _ = plot_soc(sim_df, outdir)
    _ = plot_cashflow(ab_df, outdir)
    _ = plot_price_and_dispatch(base, sim_df, outdir)

    # --- Redispatch (optional) ---
    if args.rd_csv:
        rd_df_raw = read_redispatch_csv(Path(args.rd_csv))
        calls_norm = normalize_redispatch_calls(rd_df_raw, base.copy())
        sim_base, ab_base, sim_rd, ab_rd, rd_report, base_adj, calls_norm = redispatch_analysis(base.copy(), params, calls_norm)

        # Export RD CSVs
        calls_norm.to_csv(outdir / "redispatch_calls_normalized.csv", index=False)
        rd_report.to_csv(outdir / "abrechnung_redispatch.csv", index=False)

        # Plot incremental cumulative
        fig_path = outdir / "Redispatch_Kumulierte_DeltaCF.png"
        plt.figure(figsize=(10,5))
        x = rd_report["timestamp"] if "timestamp" in rd_report.columns and rd_report["timestamp"].notna().any() else rd_report["slot"]
        plt.plot(x, rd_report["Kumuliert_ΔCF_with_Comp_[EUR]"], linewidth=2)
        plt.title("Redispatch: Kumulierte Δ-Cashflows inkl. Compensation")
        plt.xlabel("Zeit" if ('timestamp' in rd_report.columns and rd_report['timestamp'].notna().any()) else "Slot")
        plt.ylabel("EUR")
        plt.grid(True, alpha=0.3)
        plt.tight_layout(); plt.savefig(fig_path, dpi=150); plt.close()

        if args.write_excel:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            wb = load_workbook(xlsx_path)
            # write sheets via pandas
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                calls_norm.to_excel(writer, sheet_name="Redispatch_Calls", index=False)
                rd_report.to_excel(writer, sheet_name="Abrechnung_Redispatch", index=False)
            # add chart to Charts
            wb = load_workbook(xlsx_path)
            if "Charts" not in wb.sheetnames:
                wb.create_sheet("Charts")
            ws = wb["Charts"]
            img = XLImage(str(fig_path))
            img.anchor = "A50"
            ws.add_image(img)
            wb.save(xlsx_path)

    # KPIs printout
    total_rev = float(ab_df["Einnahmen_[EUR]"].sum())
    total_cost = float(ab_df["Kosten_[EUR]"].sum())
    total_cf = float(ab_df["Cashflow_[EUR]"].sum())
    e_in = float(ab_df["E_in_from_Grid_[MWh]"].sum())
    e_out = float(ab_df["E_out_to_Grid_[MWh]"].sum())
    rte = (e_out / e_in) if e_in > 0 else float("nan")

    print("=== KPIs (ohne Redispatch) ===")
    print(f"Energie Bezug (Grid->BESS): {e_in:.3f} MWh")
    print(f"Energie Abgabe (BESS->Grid): {e_out:.3f} MWh")
    print(f"Energetischer RTE (realisiert): {rte:.3f}")
    print(f"Erlöse gesamt: {total_rev:.2f} EUR")
    print(f"Kosten gesamt: {total_cost:.2f} EUR")
    print(f"Cashflow gesamt: {total_cf:.2f} EUR")
    print("Exports im outdir (inkl. RD-Reports wenn --rd_csv gesetzt wurde).")


if __name__ == "__main__":
    main()
