#!/usr/bin/env python3
"""
build_excel_from_csv.py – CSV → Excel-Builder im BESS-Format
- Erwartet CSV mit Spalten: timestamp/Zeit ODER hour, sowie price_eur_mwh, dispatch_mw
- Schreibt Excel mit Sheets: Basisdaten, Parameter, SoC_Simulation, Abrechnung_SoC, KPIs, Charts
"""

from pathlib import Path
import argparse
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage

def load_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    # normalize
    rn = {}
    for c in df.columns:
        lc = c.lower().strip()
        if lc in ("timestamp","zeit","time"):
            rn[c] = "timestamp"
        elif lc == "hour":
            rn[c] = "hour"
        elif lc in ("price_eur_mwh","preis_[eur/mwh]","price"):
            rn[c] = "price_eur_mwh"
        elif lc in ("dispatch_mw","bess_dispatch_[mw]","dispatch"):
            rn[c] = "dispatch_mw"
    df = df.rename(columns=rn)
    if "price_eur_mwh" not in df.columns or "dispatch_mw" not in df.columns:
        raise ValueError("CSV braucht Spalten price_eur_mwh und dispatch_mw (oder kompatible Namen).")
    if "hour" not in df.columns and "timestamp" not in df.columns:
        raise ValueError("CSV braucht entweder 'hour' oder 'timestamp'/'Zeit'.")
    return df

def default_params():
    return {
        "Wirkungsgrad Entladen": 0.92,
        "Wirkungsgrad Laden": 0.92,
        "Zeitschrittdauer [h]": 1.0,
        "Kapazität [MWh]": 8.0,
        "P_max_Entladen [MW]": 2.0,
        "P_max_Laden [MW]": 2.0,
        "SoC_init [%]": 50.0,
        "SoC_min [%]": 5.0,
        "SoC_max [%]": 95.0
    }

def simulate_soc(base, params):
    eta_dis = float(params["Wirkungsgrad Entladen"])
    eta_cha = float(params["Wirkungsgrad Laden"])
    dt = float(params["Zeitschrittdauer [h]"])
    cap = float(params["Kapazität [MWh]"])
    pmax_dis = float(params["P_max_Entladen [MW]"])
    pmax_cha = float(params["P_max_Laden [MW]"])
    soc_init_pct = float(params["SoC_init [%]"])
    soc_min_pct = float(params["SoC_min [%]"])
    soc_max_pct = float(params["SoC_max [%]"])

    soc_min = cap * soc_min_pct/100.0
    soc_max = cap * soc_max_pct/100.0
    soc = cap * soc_init_pct/100.0

    hours = base["hour"].to_list() if "hour" in base.columns else list(range(len(base)))
    price = base["price_eur_mwh"].to_numpy()
    dispatch_plan = base["dispatch_mw"].to_numpy()

    dispatch_limited, dispatch_feasible = [], []
    e_out_grid, e_in_grid = [], []
    soc_list, soc_pct_list, flags = [], [], []

    for d_plan in dispatch_plan:
        d_lim = min(max(d_plan, -pmax_cha), pmax_dis)
        d_feas = d_lim
        reason = "OK"
        if d_lim >= 0:
            e_out = d_lim * dt
            e_batt = e_out / eta_dis if eta_dis>0 else 1e18
            if soc - e_batt < soc_min - 1e-12:
                avail = max(soc - soc_min, 0.0)
                e_out_max = avail * eta_dis
                d_feas = max(e_out_max / dt, 0.0)
                reason = "Clip: SoC_min"
            e_out = d_feas * dt
            e_batt = e_out / eta_dis if eta_dis>0 else 0.0
            soc = max(soc - e_batt, soc_min)
            e_in = 0.0
        else:
            e_in = (-d_lim) * dt
            e_batt = e_in * eta_cha
            if soc + e_batt > soc_max + 1e-12:
                room = max(soc_max - soc, 0.0)
                e_in_max = room / eta_cha if eta_cha>0 else 0.0
                d_feas = -max(e_in_max / dt, 0.0)
                reason = "Clip: SoC_max"
            e_in = (-d_feas) * dt
            e_batt = e_in * eta_cha
            soc = min(soc + e_batt, soc_max)
            e_out = 0.0

        dispatch_limited.append(d_lim)
        dispatch_feasible.append(d_feas)
        e_out_grid.append(e_out)
        e_in_grid.append(e_in)
        soc_list.append(soc)
        soc_pct_list.append(100.0*soc/cap if cap>0 else 0.0)
        if reason=="OK" and abs(d_lim - d_plan) > 1e-12:
            reason="Clip: P_max"
        flags.append(reason)

    sim = pd.DataFrame({
        "Stunde": hours,
        "Preis_[EUR/MWh]": price,
        "Dispatch_Plan_[MW]": dispatch_plan,
        "Dispatch_Limit_[MW]": dispatch_limited,
        "Dispatch_Feasible_[MW]": dispatch_feasible,
        "E_out_to_Grid_[MWh]": e_out_grid,
        "E_in_from_Grid_[MWh]": e_in_grid,
        "SoC_[MWh]": soc_list,
        "SoC_[%]": soc_pct_list,
        "Check": flags
    })
    if "timestamp" in base.columns:
        sim["timestamp"] = pd.to_datetime(base["timestamp"])
    return sim

def settlement_from_sim(sim_df):
    price = sim_df["Preis_[EUR/MWh]"].to_numpy()
    e_out = sim_df["E_out_to_Grid_[MWh]"].to_numpy()
    e_in = sim_df["E_in_from_Grid_[MWh]"].to_numpy()
    revenue = e_out * price
    costs = e_in * price
    cash = revenue - costs
    df = pd.DataFrame({
        "Stunde": sim_df["Stunde"].to_numpy(),
        "Preis_[EUR/MWh]": price,
        "E_out_to_Grid_[MWh]": e_out,
        "E_in_from_Grid_[MWh]": e_in,
        "Einnahmen_[EUR]": revenue,
        "Kosten_[EUR]": costs,
        "Cashflow_[EUR]": cash
    })
    df["Kumuliert_[EUR]"] = df["Cashflow_[EUR]"].cumsum()
    if "timestamp" in sim_df.columns:
        df["timestamp"] = sim_df["timestamp"]
    return df

def make_charts(sim_df, ab_df, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    # SoC
    plt.figure(figsize=(10,5))
    x = sim_df["timestamp"] if "timestamp" in sim_df.columns else sim_df["Stunde"]
    plt.plot(x, sim_df["SoC_[%]"], linewidth=2)
    plt.title("SoC-Verlauf [%]"); plt.xlabel("Zeit" if "timestamp" in sim_df.columns else "Stunde"); plt.ylabel("SoC [%]")
    plt.grid(True, alpha=0.3); plt.tight_layout()
    soc_png = outdir/"SoC_Verlauf.png"; plt.savefig(soc_png, dpi=150); plt.close()

    # CF kumuliert
    plt.figure(figsize=(10,5))
    x2 = ab_df["timestamp"] if "timestamp" in ab_df.columns else ab_df["Stunde"]
    plt.plot(x2, ab_df["Kumuliert_[EUR]"], linewidth=2)
    plt.title("Kumulierter Cashflow [EUR]"); plt.xlabel("Zeit" if "timestamp" in ab_df.columns else "Stunde"); plt.ylabel("Kumuliert [EUR]")
    plt.grid(True, alpha=0.3); plt.tight_layout()
    cf_png = outdir/"Cashflow_Kumuliert.png"; plt.savefig(cf_png, dpi=150); plt.close()

    return soc_png, cf_png

def write_excel(excel_path: Path, base_df, params, sim_df, ab_df, soc_png, cf_png):
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Basisdaten
        base_out = base_df.copy()
        remap = {}
        if "hour" in base_out.columns: remap["hour"] = "Stunde"
        if "price_eur_mwh" in base_out.columns: remap["price_eur_mwh"] = "Strompreis_[EUR/MWh]"
        if "dispatch_mw" in base_out.columns: remap["dispatch_mw"] = "BESS_Dispatch_[MW]"
        base_out = base_out.rename(columns=remap)
        base_out.to_excel(writer, sheet_name="Basisdaten", index=False)

        # Parameter
        params_df = pd.DataFrame({
            "Parameter": list(params.keys()),
            "Wert": list(params.values()),
            "Hinweis": [
                "Energie an Netz = E_out; interner Bedarf = E_out/eta_dis",
                "Energie in Speicher = E_in*eta_cha",
                "Dauer pro Zeitschritt",
                "Speicherkapazität (MWh)",
                "Begrenzung Entladeleistung",
                "Begrenzung Ladeleistung",
                "Startfüllstand [%]",
                "Untergrenze SoC [%]",
                "Obergrenze SoC [%]"
            ]
        })
        params_df.to_excel(writer, sheet_name="Parameter", index=False)

        # Simulation & Abrechnung
        sim_df.to_excel(writer, sheet_name="SoC_Simulation", index=False)
        ab_df.to_excel(writer, sheet_name="Abrechnung_SoC", index=False)

        # KPIs
        kpis = pd.DataFrame({
            "KPI": [
                "Kapazität [MWh]","SoC_min [%]","SoC_max [%]","P_max_Entladen [MW]","P_max_Laden [MW]",
                "Round-Trip-Effizienz (vorgegeben)","Geladene Energie (Grid->BESS) [MWh]",
                "Entladene Energie (BESS->Grid) [MWh]","Energetischer RTE (realisiert)",
                "Erlöse gesamt [EUR]","Kosten gesamt [EUR]","Cashflow gesamt [EUR]"
            ],
            "Wert": [
                params["Kapazität [MWh]"], params["SoC_min [%]"], params["SoC_max [%]"],
                params["P_max_Entladen [MW]"], params["P_max_Laden [MW]"],
                params["Wirkungsgrad Entladen"]*params["Wirkungsgrad Laden"],
                float(ab_df["E_in_from_Grid_[MWh]"].sum()),
                float(ab_df["E_out_to_Grid_[MWh]"].sum()),
                float(ab_df["E_out_to_Grid_[MWh]"].sum()/ab_df["E_in_from_Grid_[MWh]"].sum() if ab_df["E_in_from_Grid_[MWh]"].sum()>0 else np.nan),
                float(ab_df["Einnahmen_[EUR]"].sum()),
                float(ab_df["Kosten_[EUR]"].sum()),
                float(ab_df["Cashflow_[EUR]"].sum())
            ]
        })
        kpis.to_excel(writer, sheet_name="KPIs", index=False)

    # Charts sheet
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = load_workbook(excel_path)
    if "Charts" in wb.sheetnames:
        ws = wb["Charts"]
        wb.remove(ws)
    ws = wb.create_sheet("Charts")
    ws["A1"] = "SoC-Verlauf [%]"; ws["A25"] = "Kumulierter Cashflow [EUR]"
    img1 = XLImage(str(soc_png)); img1.anchor = "A2"; ws.add_image(img1)
    img2 = XLImage(str(cf_png)); img2.anchor = "A26"; ws.add_image(img2)
    wb.save(excel_path)

def main():
    ap = argparse.ArgumentParser(description="CSV → Excel-Builder für BESS")
    ap.add_argument("--csv", required=True, help="Pfad zur CSV (timestamp/hour, price_eur_mwh, dispatch_mw)")
    ap.add_argument("--excel", required=True, help="Zielpfad zur Excel-Datei")
    ap.add_argument("--eta_dis", type=float)
    ap.add_argument("--eta_cha", type=float)
    ap.add_argument("--dt", type=float)
    ap.add_argument("--cap", type=float)
    ap.add_argument("--pmax_dis", type=float)
    ap.add_argument("--pmax_cha", type=float)
    ap.add_argument("--soc_init", type=float)
    ap.add_argument("--soc_min", type=float)
    ap.add_argument("--soc_max", type=float)
    args = ap.parse_args()

    csv_path = Path(args.csv).expanduser().resolve()
    xlsx_path = Path(args.excel).expanduser().resolve()

    base_df = load_csv(csv_path)
    if "hour" not in base_df.columns and "timestamp" in base_df.columns:
        base_df["hour"] = range(len(base_df))

    params = default_params()
    if args.eta_dis is not None: params["Wirkungsgrad Entladen"] = args.eta_dis
    if args.eta_cha is not None: params["Wirkungsgrad Laden"] = args.eta_cha
    if args.dt is not None: params["Zeitschrittdauer [h]"] = args.dt
    if args.cap is not None: params["Kapazität [MWh]"] = args.cap
    if args.pmax_dis is not None: params["P_max_Entladen [MW]"] = args.pmax_dis
    if args.pmax_cha is not None: params["P_max_Laden [MW]"] = args.pmax_cha
    if args.soc_init is not None: params["SoC_init [%]"] = args.soc_init
    if args.soc_min is not None: params["SoC_min [%]"] = args.soc_min
    if args.soc_max is not None: params["SoC_max [%]"] = args.soc_max

    sim_df = simulate_soc(base_df, params)
    ab_df = settlement_from_sim(sim_df)

    outdir = xlsx_path.parent / (xlsx_path.stem + "_charts")
    soc_png, cf_png = make_charts(sim_df, ab_df, outdir)

    write_excel(xlsx_path, base_df, params, sim_df, ab_df, soc_png, cf_png)
    print("Fertig gebaut:", xlsx_path)
    print("Charts:", outdir)

if __name__ == "__main__":
    main()
